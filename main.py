# import library
from tkinter import *
from tkinter.ttk import Combobox
import tkinter as tk
from tkinter import messagebox
import openpyxl
from openpyxl import Workbook
import pathlib
from queue import Queue
from collections import deque

# Stack untuk menyimpan data yang telah disubmit
submission_stack = deque()

# Queue untuk menyimpan data yang di-clear dengan batas maksimal 10 item
clearance_queue = Queue(maxsize=10)

# Mengecek apakah file Excel sudah ada; jika belum, buat dengan header
file = pathlib.Path('Backened_data.xlsx')

if file.exists():
    pass
else:
    file = Workbook()
    sheet = file.active
    sheet['A1'] = "Full Name"
    sheet['B1'] = "PhoneNumber"
    sheet['C1'] = "Age"
    sheet['D1'] = "Gender"
    sheet['E1'] = "Address"
    
    file.save('Backened_data.xlsx')

# Fungsi untuk menyimpan data ke Excel dan stack
def submit():
    name = nameValue.get()
    contact = contactValue.get()
    age = AgeValue.get()
    gender = gender_combobox.get()
    address = addressEntry.get(1.0, END)

    # Simpan data yang disubmit ke dalam stack
    submission_stack.append((name, contact, age, gender, address))

    # Membuka file Excel dan menambah data baru ke baris berikutnya
    file = openpyxl.load_workbook('Backened_Data.xlsx')
    sheet = file.active
    sheet.cell(column=1, row=sheet.max_row+1, value=name)
    sheet.cell(column=2, row=sheet.max_row, value=contact)
    sheet.cell(column=3, row=sheet.max_row, value=age)
    sheet.cell(column=4, row=sheet.max_row, value=gender)
    sheet.cell(column=5, row=sheet.max_row, value=address)

    file.save('Backened_data.xlsx')

    # Menampilkan popup riwayat terakhir
    messagebox.showinfo("Submission History", f"Last Submitted:\nName: {name}\nContact: {contact}\nAge: {age}\nGender: {gender}\nAddress: {address}")

# Fungsi untuk meng-clear data dari form dan menyimpannya ke dalam queue
def clear():
    name = nameValue.get()
    contact = contactValue.get()
    age = AgeValue.get()
    gender = gender_combobox.get()
    address = addressEntry.get(1.0, END)

    # Simpan data yang di-clear ke dalam queue
    clearance_queue.put((name, contact, age, gender, address))

    # Hapus isi form
    nameValue.set('')
    contactValue.set('')
    AgeValue.set('')
    addressEntry.delete(1.0, END)

    # Menampilkan popup riwayat terakhir
    messagebox.showinfo("Clearance History", f"Last Cleared:\nName: {name}\nContact: {contact}\nAge: {age}\nGender: {gender}\nAddress: {address}")

# Setup jendela utama aplikasi
root = Tk()
root.title("Data Entry")
root.geometry('700x400+300+200')
root.resizable(False, False)
root.configure(bg="#326273")

# Icon aplikasi
icon_image = PhotoImage(file="Flyaway.png")
root.iconphoto(False, icon_image)

# Heading aplikasi
Label(root, text="Please fill out this Entry form:", font="arial 13", bg="#326273", fg="#fff").place(x=20, y=20)

# Label untuk setiap field
Label(root, text='Name', font=23, bg="#326273", fg="#fff").place(x=50, y=100)
Label(root, text='Contact No.', font=23, bg="#326273", fg="#fff").place(x=50, y=150)
Label(root, text='Age', font=23, bg="#326273", fg="#fff").place(x=50, y=200)
Label(root, text='Gender', font=23, bg="#326273", fg="#fff").place(x=350, y=200)
Label(root, text='Address', font=23, bg="#326273", fg="#fff").place(x=50, y=250)

# Variabel untuk menyimpan nilai dari entry fields
nameValue = StringVar()
contactValue = StringVar()
AgeValue = StringVar()

# Entry fields untuk nama, kontak, dan usia
nameEntry = Entry(root, textvariable=nameValue, width=37, bd=2, font=20)
contactEntry = Entry(root, textvariable=contactValue, width=37, bd=2, font=20)
ageEntry = Entry(root, textvariable=AgeValue, width=12, bd=2, font=20)

# Combobox untuk memilih gender
gender_combobox = Combobox(root, values=['Male', 'Female'], font='arial 14', state='readonly', width=13)
gender_combobox.place(x=440, y=200)
gender_combobox.set('Male')

# Text field untuk alamat
addressEntry = Text(root, width=50, height=4, bd=4)

# Menempatkan entry fields di window
nameEntry.place(x=200, y=100)
contactEntry.place(x=200, y=150)
ageEntry.place(x=200, y=200)
addressEntry.place(x=200, y=250)

# Tombol untuk submit, clear, dan exit
Button(root, text="Submit", bg="#326273", fg="white", width=15, height=2, command=submit).place(x=200, y=350)
Button(root, text="Clear", bg="#326273", fg="white", width=15, height=2, command=clear).place(x=340, y=350)
Button(root, text="Exit", bg="#326273", fg="white", width=15, height=2, command=lambda: root.destroy()).place(x=480, y=350)

# Menampilkan isi stack submission
print("Submission Stack:")
for data in submission_stack:
    print(data)

# Menampilkan isi queue clearance
print("\nClearance Queue:")
while not clearance_queue.empty():
    print(clearance_queue.get())

# Menjalankan aplikasi Tkinter
root.mainloop()
